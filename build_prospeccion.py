from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill('solid', fgColor='2F5496')
HEADER_FONT = Font(bold=True, color='FFFFFF')
TITLE_FONT = Font(bold=True, size=13, color='2F5496')
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

VOL_FILLS = {
    'Alto': PatternFill('solid', fgColor='C6EFCE'),
    'Medio': PatternFill('solid', fgColor='FFEB9C'),
    'Bajo': PatternFill('solid', fgColor='FFC7CE'),
    'Bajo-Medio': PatternFill('solid', fgColor='FFEB9C'),
}
VOL_FONTS = {
    'Alto': Font(color='006100'),
    'Medio': Font(color='9C6500'),
    'Bajo': Font(color='9C0006'),
    'Bajo-Medio': Font(color='9C6500'),
}

TRACK_HEADERS = ['Fecha de contacto', 'Resultado de llamada', 'Próxima cita/acción']
TRACK_WIDTHS = [18, 32, 24]
TRACK_FILL = PatternFill('solid', fgColor='FFE699')
TRACK_FONT = Font(bold=True, color='7F6000')

def style_header(ws, row, ncols, track_count=0):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        if track_count and c > ncols - track_count:
            cell.fill = TRACK_FILL
            cell.font = TRACK_FONT
        else:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_table(ws, start_row, headers, rows, vol_col=None, widths=None, track_count=0):
    if track_count:
        headers = list(headers) + TRACK_HEADERS[:track_count]
        rows = [list(r) + [''] * track_count for r in rows]
        widths = (list(widths) + TRACK_WIDTHS[:track_count]) if widths else widths
    style_header(ws, start_row, len(headers), track_count=track_count)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=c, value=h)
    for r, row_data in enumerate(rows, start=start_row + 1):
        for c, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if vol_col is not None and c == vol_col and val in VOL_FILLS:
                cell.fill = VOL_FILLS[val]
                cell.font = VOL_FONTS[val]
    if widths:
        autosize(ws, widths)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

wb = Workbook()

# HOJA 1 - Resumen
ws = wb.active
ws.title = 'Resumen'
ws['A1'] = 'Plan de Prospección B2B — El Buen Sazón (Abarrotes y Perecederos)'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:E1')
headers = ['Tipo de cliente', 'Volumen de compra estimado', 'Prioridad (1-5)', 'Estatus general', 'Próxima acción']
rows = [
    ['Gobierno del Estado (licitaciones/cotización)', 'Alto', 1, 'Listos para cotizar (ya tienen REPSE y venta previa a gobierno)', 'Subir cotización en portal de Investigación de Mercado'],
    ['DIF / Casas de asistencia', 'Alto', 1, 'En seguimiento de licitaciones', 'Monitorear portal mensualmente'],
    ['Comedores industriales (Eurest/Compass, Sodexo)', 'Alto', 2, 'Contacto inicial pendiente', 'Llamar a matriz para contacto Puebla'],
    ['Asilos / residencias adultos mayores', 'Medio', 3, 'Contactos identificados', 'Llamar directo'],
    ['Escuelas privadas con comedor', 'Medio', 4, 'Pendiente de listado completo', 'Recorrer directorio por zona'],
    ['Restaurantes independientes', 'Bajo-Medio', 5, 'Listado inicial armado', 'Prospección telefónica'],
]
write_table(ws, 3, headers, rows, vol_col=2, widths=[42, 16, 14, 40, 38])

# HOJA 2 - Gobierno Licitaciones
ws = wb.create_sheet('Gobierno - Licitaciones')
ws['A1'] = 'Licitaciones de Gobierno del Estado de Puebla — Insumos Alimentarios'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:F1')
headers = ['Número de licitación', 'Dependencia', 'Objeto', 'Fecha publicación', 'Estatus', 'Notas']
rows = [
    ['GEP-SPFA-LPN-102-127/2026', 'DIF Estatal (SEDIF)', 'Adquisición de insumos alimentarios: abarrotes, carnes rojas, pollo, pescado, lácteos, embutidos, frutas, verduras y tortillas', 'Abril 2026', 'Cerrada', 'Casas de Asistencia'],
    ['GEP-SPFA-LPN-058-127/2026', 'DIF Estatal (SEDIF)', 'Mismo giro de insumos alimentarios', 'Marzo 2026', 'Cerrada', 'Casas de Asistencia'],
    ['GEP-SPFA-LPN-104-085/2026', 'ISSSTEP', 'Servicio integral de abasto y distribución de alimentos perecederos', '04/05/2026', 'Cerrada', '-'],
    ['GEP-SPFA-LPN-135-060/2026', 'Convenciones y Parques', 'Adquisición de insumos alimenticios perecederos, no perecederos y cárnicos', '28/05/2026', 'Cerrada', '-'],
    ['GEP-SPFA-LPN-052-209/2026', 'Secretaría de Seguridad Pública', 'Insumos, víveres, productos alimenticios y tortillas perecederos y no perecederos para centros penitenciarios', 'Marzo 2026', 'Cerrada', '-'],
    ['GEP-SPFA-LPN-053-211/2026', 'Secretaría de Seguridad Pública', 'Paquetes de insumos perecederos y no perecederos para comedor', 'Marzo 2026', 'Cerrada', '-'],
    ['GEP-SPFA-LPN-003-014/2026', 'Universidad / Seguridad Pública', 'Abasto de alimentos para elementos en pernocta y capacitación', 'Enero 2026', 'Cerrada', '-'],
    ['(próxima convocatoria)', 'Por monitorear', 'Pendiente de nueva publicación', '-', 'Pendiente', 'Revisar https://licitaciones.puebla.gob.mx mensualmente'],
]
write_table(ws, 3, headers, rows, widths=[26, 26, 55, 16, 12, 40])

note_row = 3 + len(rows) + 2
ws.cell(row=note_row, column=1, value='IMPORTANTE:').font = Font(bold=True, color='9C0006')
note_cell = ws.cell(row=note_row + 1, column=1,
    value='Enviar cotización permanente en el módulo de Investigación de Mercado: '
          'http://cga.sfapuebla.gob.mx/SISDABS/Publico/Inicio/Cotizaciones — esto permite quedar en '
          'consideración para adjudicaciones directas sin esperar una LPN abierta.')
note_cell.font = Font(italic=True, color='9C0006')
note_cell.alignment = Alignment(wrap_text=True)
ws.merge_cells(start_row=note_row + 1, start_column=1, end_row=note_row + 1, end_column=6)
ws.row_dimensions[note_row + 1].height = 40

# HOJA 3 - Comedores Industriales
ws = wb.create_sheet('Comedores Industriales')
ws['A1'] = 'Concesionarios de Comedores Industriales'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:I1')
headers = ['Empresa', 'Tipo', 'Contacto/Correo', 'Teléfono u oficina', 'Zona objetivo', 'Notas']
rows = [
    ['Eurest / Compass Group México', 'Concesionario de comedores industriales', 'contacto@eurest.com.mx (formato nombre.apellido@eurest.com.mx)', 'Matriz CDMX, Av. Jaime Balmes 11 Torre D Polanco', 'Parques industriales Puebla (VW, Audi, Faurecia, Hella)', 'Pedir que canalicen a responsable de compras/abasto Puebla'],
    ['Sodexo México', 'Concesionario de comedores industriales', '-', '-', 'Parques industriales Puebla', 'Buscar contacto directo'],
    ['Newrest Catering México', 'Concesionario de comedores industriales', '-', '-', 'Parques industriales Puebla', 'Buscar contacto directo'],
]
write_table(ws, 3, headers, rows, widths=[30, 30, 42, 38, 36, 40], track_count=3)

# HOJA 4 - Asilos y Residencias
ws = wb.create_sheet('Asilos y Residencias')
ws['A1'] = 'Asilos y Residencias de Adultos Mayores'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:H1')
headers = ['Nombre', 'Dirección', 'Teléfono', 'Volumen estimado', 'Notas']
rows = [
    ['El Hogar de mi Abuelito (suc. Jardines de San Manuel)', 'Río Pánuco 5331, Jardines de San Manuel, Puebla', '222-213-2360', 'Medio', '-'],
    ['El Hogar de mi Abuelito (suc. Bugambilias)', 'Orquídeas 6327, Bugambilias, Puebla', '222-228-9648', 'Medio', '-'],
    ['Casa de Descanso La Tercera Edad', 'Calle 2-B Sur 5914, Col. Bugambilias, Puebla', '-', 'Medio', 'Más de 20 años de experiencia'],
    ['Residencia Consuelo Gutiérrez', 'Puebla', '-', 'Medio', 'Especializada en Alzheimer y demencias'],
    ['Fundación Alborada', 'Puebla', '-', 'Medio', 'Buscar teléfono directo'],
]
write_table(ws, 3, headers, rows, vol_col=4, widths=[42, 42, 16, 16, 38], track_count=3)

# HOJA 5 - Restaurantes
ws = wb.create_sheet('Restaurantes')
ws['A1'] = 'Restaurantes Independientes y Cadenas'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:I1')
headers = ['Nombre', 'Tipo de cocina', 'Dirección', 'Sucursales', 'Volumen estimado', 'Notas']
rows = [
    ['Antigua Taquería La Oriental', 'Taquería', 'Blvd. Héroes del 5 de Mayo 3126, Ladrillera de Benítez', 42, 'Alto', 'Cadena grande, alto volumen'],
    ['Mr. Pampas', 'Brasileña/Buffet', 'Av. Juárez 2924-A, La Paz', 18, 'Alto', 'Cadena grande'],
    ['Restaurante Cabo San Lucas', 'Mariscos', 'Av. Las Torres 2602, Tres Cruces', 3, 'Medio', '-'],
    ['La Piccola Nostra', 'Italo-mexicana', 'Teziutlán Norte 1, La Paz', 1, 'Bajo-Medio', '-'],
    ['El Burladero Las Ánimas', 'Típica/Poblana/Carnes', 'Sta. Anita Nte. Lt 1, Estrellas del Sur', 1, 'Bajo-Medio', '-'],
    ['Cocina Típica Mexicana Lupita', 'Mexicana', '4 Poniente 1106, Centro', 1, 'Bajo', '-'],
    ['La Encomienda', 'Cortes/Internacional', 'Blvd. Atlixcáyotl 1499, San Andrés Cholula', 1, 'Medio', '-'],
    ['Buffalucas', 'Hamburguesas', '39 Oriente 1407 local 6', 1, 'Bajo', '-'],
    ['Sacristía de Capuchinas', 'Restaurante', '9 Oriente 16, Centro', 1, 'Medio', '-'],
]
write_table(ws, 3, headers, rows, vol_col=5, widths=[32, 24, 42, 12, 16, 32], track_count=3)

# HOJA 6 - Escuelas Privadas (pendiente)
ws = wb.create_sheet('Escuelas Privadas (pendiente)')
ws['A1'] = 'Escuelas Privadas — Pendiente de Levantamiento'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:H1')
headers = ['Nombre', 'Dirección', 'Nivel educativo', 'Teléfono', 'Notas']
rows = [
    ['-', '-', '-', '-', 'Pendiente de recorrer directorio TODOPUEBLA (427 escuelas registradas) por zona'],
]
write_table(ws, 3, headers, rows, widths=[32, 32, 22, 16, 50], track_count=3)
ref_row = 3 + len(rows) + 2
ref_cell = ws.cell(row=ref_row, column=1, value='Referencia: https://www.todopuebla.com/directorio/escuelasyguarderias/escuelasprivadas')
ref_cell.font = Font(italic=True, color='2F5496')
ws.merge_cells(start_row=ref_row, start_column=1, end_row=ref_row, end_column=8)

wb.save('/root/prospeccion_el_buen_sazon.xlsx')
print('OK')
